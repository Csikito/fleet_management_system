import csv
from io import BytesIO, StringIO

from flask import send_file

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from .util import get_vehicle_type_status_name, get_vehicle_model_status_name



def export_csv(vehicles):
    output = StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow(["ID", "License Plate", "Model", "Year", "Engine Type", "Registration Expiry Date"])

    # Data
    for v in vehicles:
        engine_type = get_vehicle_type_status_name(True).get(v.type, "-"),
        model = get_vehicle_model_status_name(True).get(v.model, "-"),

        writer.writerow([v.id, v.license_plate, model, v.year, engine_type, v.registration_expiry_date])

    # StringIO to bytes
    output_bytes = BytesIO(output.getvalue().encode("utf-8"))
    output_bytes.seek(0)

    return send_file(
        output_bytes,
        mimetype="text/csv",
        as_attachment=True,
        download_name=f"vehicle_report.csv"
    )


def export_xlsx(vehicles):
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Vehicle Report"

    # Define column headers
    headers = ["Type", "Model", "Year", "License Plate", "Mileage", "Reg. Expiry"]
    ws.append(headers)

    # Apply styles to header row
    header_font = Font(bold=True, color="FFFFFF")  # White text
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Blue background
    center_alignment = Alignment(horizontal="center", vertical="center")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    # Populate the Excel sheet with vehicle data
    for i, v in enumerate(vehicles, start=2):  # Start from row 2 (row 1 is the header)
        vehicles_type = get_vehicle_type_status_name(True).get(v.type, "-")
        model = get_vehicle_model_status_name(True).get(v.model, "-")
        registration_expiry_date = str(v.registration_expiry_date).replace("-", ".")

        row = [vehicles_type, model, v.year, v.license_plate, v.current_mileage, registration_expiry_date]
        ws.append(row)

        # Apply alternating row color (light gray for even rows)
        if i % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=i, column=col).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9",
                                                              fill_type="solid")

    # Auto-adjust column widths based on content
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Get the column letter (e.g., A, B, C)
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2  # Adjust width

    # Save the workbook and send the file
    wb.save(output)
    output.seek(0)

    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"vehicle_report.xlsx")


def export_pdf(vehicles):
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)

    # TITLE STYLE (centered, larger font size)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontSize=18,
        alignment=1,  # 0=left, 1=center, 2=right alignment
        spaceAfter=20,  # Bottom margin (spacing from the table)
    )

    # ADD TITLE
    title = Paragraph("Vehicle Report", title_style)

    # TABLE HEADER
    data = [
        ["Type", "Model", "Year", "License Plate", "Mileage", "Reg. Expiry"]
    ]

    # ADD DATA (every second row will have a light gray background)
    for i, v in enumerate(vehicles, start=1):
        vehicles_type = get_vehicle_type_status_name(True).get(v.type, "-")
        model = get_vehicle_model_status_name(True).get(v.model, "-")
        registration_expiry_date = str(v.registration_expiry_date).replace("-", ".")

        row = [vehicles_type, model, v.year, v.license_plate, v.current_mileage, registration_expiry_date]
        data.append(row)

    # CREATE TABLE
    table = Table(data, colWidths=[80, 100, 50, 100, 80, 80])

    # TABLE STYLES
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Header background color
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),  # Header text color (white)
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Center text alignment
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Header font
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),  # Data font
        ('FONTSIZE', (0, 0), (-1, 0), 12),  # Header font size
        ('FONTSIZE', (0, 1), (-1, -1), 10),  # Data font size
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),  # Header padding
        ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Table borders
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),  # Default row background color
    ])

    # Apply alternating row background color (light gray every second row)
    for i in range(1, len(data)):
        if i % 2 == 0:
            style.add('BACKGROUND', (0, i), (-1, i), colors.lightgrey)

    table.setStyle(style)

    # BUILD DOCUMENT STRUCTURE
    elements = [title, Spacer(1, 20), table]  # Spacer = empty space between title and table
    doc.build(elements)

    output.seek(0)
    return send_file(output, mimetype="application/pdf", as_attachment=True,
                     download_name=f"vehicle_report.pdf")
